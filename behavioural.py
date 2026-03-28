"""
Behavioural Statistical Analysis — Article 1 (Control vs PTSD)
Boxplot + jitter, ggplot2 style. Separate figures for OF, EPM, DLB.

Statistical pipeline:
    1. Outlier removal — IQR 1.5×
    2. Normality — Shapiro-Wilk per group
    3. Overall — one-way ANOVA (all normal) or Kruskal-Wallis
    4. Post-hoc — t-test (after ANOVA) or Mann-Whitney U (after Kruskal)

Usage:
    python behav_art1.py
    python behav_art1.py --test EPM
    python behav_art1.py --no-outliers

Data files in data/ folder:
    data/OF_DF_ALL.xlsx
    data/EP_DF_ALL.xlsx
    data/DL_DF_ALL.xlsx
"""

import os, csv, argparse
from datetime import datetime
from collections import defaultdict

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ── Config ──────────────────────────────────────────────────────────────────────

DATA_FILES = {
    'OF':  'data/OF_DF_ALL.xlsx',
    'EPM': 'data/EP_DF_ALL.xlsx',
    'DLB': 'data/DL_DF_ALL.xlsx',
}

EPM_MAP = {'cont': 'control', 'ptsd': 'ptsd'}

GROUPS = {
    'OF':  {'Control': 'control_non', 'PTSD': 'ptsd_non'},
    'EPM': {'Control': 'control_non', 'PTSD': 'ptsd_non'},
    'DLB': {'Control': 'cont_non',    'PTSD': 'ptsd_non'},
}

PARAMS = {
    'OF': {
        'distance':          'Distance (m)',
        'mean_speed':        'Speed (m/s)',
        'freezing_episodes': 'Freezing episodes (n)',
        'time_freezing':     'Freezing time (s)',
        'center_entries':    'Center entries (n)',
        'center_time':       'Center time (s)',
        'corners_entries':   'Corner entries (n)',
        'corners_time':      'Corner time (s)',
        'sides_entries':     'Side entries (n)',
        'sides_time':        'Side time (s)',
    },
    'EPM': {
        'distance':          'Distance (m)',
        'time_freezing':     'Freezing time (s)',
        'freezing_episodes': 'Freezing episodes (n)',
        'open_entries':      'Open arm entries (n)',
        'open_time':         'Open arm time (s)',
        'open_head_entries': 'Open head dips (n)',
        'closed_entries':    'Closed arm entries (n)',
        'closed_time':       'Closed arm time (s)',
        'rotations':         'Rotations (n)',
    },
    'DLB': {
        'entries':   'Light zone entries (n)',
        'total_out': 'Time in light (s)',
        'total_in':  'Time in dark (s)',
        'curiosity': 'Curiosity (n)',
    },
}

CTRL_COLOR = '#388E3C'
PTSD_COLOR = '#7B1FA2'

plt.rcParams.update({
    'font.family':  'DejaVu Sans',
    'pdf.fonttype': 42,
    'ps.fonttype':  42,
})

# ── Data loading ────────────────────────────────────────────────────────────────

def load_data(filepath, test_name):
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    cols = [c.value for c in ws[1]]
    data = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        g = str(row[1]).lower()
        t = str(row[2]).lower()
        if test_name == 'EPM' and g in EPM_MAP:
            g = EPM_MAP[g]
        data[f'{g}_{t}'].append(row)
    return cols, data


def get_vals(data, cols, group_key, param):
    if param not in cols: return []
    ci = cols.index(param)
    return [float(r[ci]) for r in data.get(group_key, [])
            if r[ci] is not None and isinstance(r[ci], (int, float))]


def remove_iqr(vals, factor=1.5):
    if len(vals) < 4: return vals
    q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
    iqr = q3 - q1
    return [v for v in vals if q1 - factor*iqr <= v <= q3 + factor*iqr]


# ── Statistics ──────────────────────────────────────────────────────────────────

def shapiro_wilk(v):
    if len(v) < 3: return False, None
    _, p = stats.shapiro(v)
    return p >= 0.05, round(p, 4)


def run_stats(groups: dict) -> dict:
    normality = {}
    for name, vals in groups.items():
        is_norm, p_sw = shapiro_wilk(vals)
        normality[name] = {'normal': is_norm, 'p': p_sw}

    all_normal = all(v['normal'] for v in normality.values()
                     if v['p'] is not None)
    values_list = list(groups.values())

    if all_normal:
        stat, p_ov = stats.f_oneway(*values_list)
        ov_test = 'One-way ANOVA'
    else:
        stat, p_ov = stats.kruskal(*values_list)
        ov_test = 'Kruskal-Wallis'

    posthoc = {}
    names = list(groups.keys())
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            g1, g2 = names[i], names[j]
            v1, v2 = groups[g1], groups[g2]
            if len(v1) < 3 or len(v2) < 3: continue
            if all_normal:
                _, p = stats.ttest_ind(v1, v2)
                ph_test = 't-test'
            else:
                _, p = stats.mannwhitneyu(v1, v2, alternative='two-sided')
                ph_test = 'Mann-Whitney U'
            p = round(p, 6)
            sl = sig_label(p)
            posthoc[(g1, g2)] = {
                'test': ph_test, 'p': p, 'sig': sl,
                'direction': (f'{g2} ↑' if np.mean(v2) > np.mean(v1)
                              else f'{g2} ↓') if sl not in ('ns', '') else '—'
            }

    return {
        'normality':  normality,
        'all_normal': all_normal,
        'overall':    {'test': ov_test, 'statistic': round(stat, 4),
                       'p': round(p_ov, 6)},
        'posthoc':    posthoc,
    }


def sig_label(p):
    if p is None: return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'


def print_stats(param_label, sr, n_rm1=0, n_rm2=0):
    print(f"\n  {param_label}:")
    if n_rm1 or n_rm2:
        print(f"    ⚠ Outliers removed: Control={n_rm1}, PTSD={n_rm2}")
    print(f"    Normality (Shapiro-Wilk):")
    for grp, v in sr['normality'].items():
        status = '✅ normal' if v['normal'] else '❌ non-normal'
        p_str  = f"p={v['p']}" if v['p'] is not None else 'n/a'
        print(f"      {grp:<12} {p_str:<12} {status}")
    ov = sr['overall']
    print(f"    Overall: {ov['test']}  stat={ov['statistic']}  p={ov['p']}")
    ph_name = 't-test' if sr['all_normal'] else 'Mann-Whitney U'
    print(f"    Post-hoc ({ph_name}):")
    for (g1, g2), v in sr['posthoc'].items():
        marker = ' ←' if v['sig'] not in ('ns', '') else ''
        print(f"      {g2} vs {g1:<12} p={v['p']:<10} {v['sig']:<4}{marker}")


# ── Plot ─────────────────────────────────────────────────────────────────────────

def plot_param(ax, v1, v2, label, posthoc, show_ylabel=False):
    bp = ax.boxplot(
        [v1, v2],
        positions=[0, 1],
        widths=0.45,
        patch_artist=True,
        notch=False,
        medianprops=dict(color='#212121', linewidth=2.0),
        whiskerprops=dict(color='#424242', linewidth=1.2),
        capprops=dict(color='#424242', linewidth=1.2),
        flierprops=dict(marker='', markersize=0),
        boxprops=dict(linewidth=1.2),
        zorder=3,
    )
    for patch, color in zip(bp['boxes'], [CTRL_COLOR, PTSD_COLOR]):
        patch.set_facecolor(color)
        patch.set_alpha(0.55)
        patch.set_edgecolor('#424242')

    np.random.seed(42)
    for xi, vals in [(0, v1), (1, v2)]:
        jitter = np.random.uniform(-0.13, 0.13, len(vals))
        ax.scatter(xi + jitter, vals,
                   color='#212121', s=28, alpha=0.75,
                   linewidths=0, zorder=5)

    # n= labels
    ax.text(0, 0, f'n={len(v1)}', ha='center', va='top', fontsize=10,
            color='#757575', transform=ax.get_xaxis_transform())
    ax.text(1, 0, f'n={len(v2)}', ha='center', va='top', fontsize=10,
            color='#757575', transform=ax.get_xaxis_transform())

    # Y limits
    data_max = max(v1 + v2)
    top = data_max * 1.35
    bottom = -1 if min(v1 + v2) >= 0 else min(v1 + v2) * 1.1
    ax.set_ylim(-top * 0.08, top)
    ax.set_xlim(-0.6, 1.6)

    # Significance bracket
    ph = posthoc.get(('Control', 'PTSD'), {})
    sl = ph.get('sig', '')
    if sl not in ('ns', ''):
        y = top * 0.84
        h = top * 0.04
        ax.plot([0, 0, 1, 1], [y, y+h, y+h, y],
                lw=1.2, color='#212121', clip_on=False)
        ax.text(0.5, y + h*1.1, sl, ha='center', va='bottom',
                fontsize=14, color='#212121', fontweight='bold', clip_on=False)

    # Formatting
    ax.set_xticks([0, 1])
    ax.set_xticklabels(['Control', 'PTSD'],
                       fontsize=13, fontweight='bold', color='#212121')
    ax.set_title(label, fontsize=13, fontweight='bold', color='#212121', pad=6)
    if show_ylabel:
        ax.set_ylabel('Mean', fontsize=12, fontweight='bold',
                      color='#424242', labelpad=18)

    # ggplot2 style
    ax.set_facecolor('#EBEBEB')
    ax.yaxis.grid(True, color='white', linewidth=1.0, zorder=0)
    ax.xaxis.grid(False)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#BDBDBD')
    ax.spines['bottom'].set_linewidth(0.8)
    ax.tick_params(axis='x', length=0, pad=18)
    ax.tick_params(axis='y', length=0, labelsize=12)


def make_figure(test_name, param_data, stats_dict, output_prefix):
    params = PARAMS[test_name]
    n      = len(params)

    if n <= 5:
        nrows, ncols = 1, n
        figsize = (n * 4.2, 5.5)
    else:
        ncols = 5
        nrows = -(-n // ncols)
        figsize = (ncols * 4.2, nrows * 5.5)

    fig, axes = plt.subplots(nrows, ncols, figsize=figsize)
    fig.patch.set_facecolor('white')
    axes_flat = np.array(axes).flatten() if n > 1 else [axes]

    for i, (param, label) in enumerate(params.items()):
        ax = axes_flat[i]
        if param not in param_data:
            ax.set_visible(False)
            continue
        v1 = param_data[param]['ctrl']
        v2 = param_data[param]['ptsd']
        ph = stats_dict[param]['posthoc']
        plot_param(ax, v1, v2, label, ph, show_ylabel=(i % ncols == 0))

    for idx in range(n, len(axes_flat)):
        axes_flat[idx].set_visible(False)

    patches = [
        mpatches.Patch(facecolor=CTRL_COLOR, edgecolor='white', label='Control'),
        mpatches.Patch(facecolor=PTSD_COLOR, edgecolor='white', label='PTSD'),
    ]
    fig.legend(handles=patches, loc='lower center', ncol=2,
               fontsize=13, frameon=True, fancybox=False, edgecolor='#E0E0E0',
               bbox_to_anchor=(0.5, -0.04), handlelength=1.5, handleheight=1.0)

    titles = {'OF': 'Open Field Test', 'EPM': 'Elevated Plus Maze',
              'DLB': 'Dark-Light Box'}
    fig.suptitle(f'{titles[test_name]} — Control vs PTSD',
                 fontsize=13, fontweight='bold', color='#212121', y=1.02)
    fig.text(0.5, -0.08,
             'Boxplot: median, IQR, whiskers = 1.5×IQR. '
             'Dots = individual animals (outliers removed, IQR 1.5×). '
             'Normality: Shapiro–Wilk. Overall: ANOVA or Kruskal–Wallis. '
             'Post-hoc: t-test or Mann–Whitney U. * p<0.05, ** p<0.01, *** p<0.001.',
             ha='center', fontsize=8, color='#9E9E9E', style='italic')

    plt.tight_layout(rect=[0, 0.06, 1, 1])
    plt.subplots_adjust(wspace=0.38, hspace=0.55)
    fig.savefig(f'{output_prefix}.png', dpi=300,
                bbox_inches='tight', facecolor='white')
    fig.savefig(f'{output_prefix}.pdf', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f"  Figure → {output_prefix}.png / .pdf")


def save_csv(rows, filepath):
    if not rows: return
    fields = list(rows[0].keys())
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"  CSV    → {filepath}")


# ── Main ─────────────────────────────────────────────────────────────────────────

def run_test(test_name, out_dir, apply_iqr=True):
    filepath = DATA_FILES[test_name]
    params   = PARAMS[test_name]

    print(f"\n{'='*60}")
    print(f"  {test_name}  |  Outlier removal: {'ON (IQR 1.5×)' if apply_iqr else 'OFF'}")
    print(f"{'='*60}")

    cols, data = load_data(filepath, test_name)

    param_data = {}
    stats_dict = {}
    csv_rows   = []

    for param, label in params.items():
        raw1 = get_vals(data, cols, GROUPS[test_name]['Control'], param)
        raw2 = get_vals(data, cols, GROUPS[test_name]['PTSD'],    param)

        if not raw1 or not raw2:
            continue

        v1 = remove_iqr(raw1) if apply_iqr else raw1
        v2 = remove_iqr(raw2) if apply_iqr else raw2

        n_rm1 = len(raw1) - len(v1)
        n_rm2 = len(raw2) - len(v2)

        param_data[param] = {'ctrl': v1, 'ptsd': v2}

        sr = run_stats({'Control': v1, 'PTSD': v2})
        stats_dict[param] = sr

        print_stats(label, sr, n_rm1, n_rm2)

        ph = sr['posthoc'].get(('Control', 'PTSD'), {})
        m1, m2 = np.mean(v1), np.mean(v2)
        s1, s2 = np.std(v1, ddof=1), np.std(v2, ddof=1)

        csv_rows.append({
            'test':                   test_name,
            'parameter':              label,
            'normality_ctrl_p':       sr['normality']['Control']['p'],
            'normality_ctrl':         'normal' if sr['normality']['Control']['normal'] else 'non-normal',
            'normality_ptsd_p':       sr['normality']['PTSD']['p'],
            'normality_ptsd':         'normal' if sr['normality']['PTSD']['normal'] else 'non-normal',
            'overall_test':           sr['overall']['test'],
            'overall_statistic':      sr['overall']['statistic'],
            'overall_p':              sr['overall']['p'],
            'posthoc_test':           ph.get('test', ''),
            'posthoc_p':              ph.get('p', ''),
            'significance':           ph.get('sig', ''),
            'direction':              ph.get('direction', ''),
            'mean_ctrl':              round(m1, 4),
            'sd_ctrl':                round(s1, 4),
            'sem_ctrl':               round(s1/np.sqrt(len(v1)), 4),
            'n_ctrl':                 len(v1),
            'mean_ptsd':              round(m2, 4),
            'sd_ptsd':                round(s2, 4),
            'sem_ptsd':               round(s2/np.sqrt(len(v2)), 4),
            'n_ptsd':                 len(v2),
            'outliers_removed_ctrl':  n_rm1,
            'outliers_removed_ptsd':  n_rm2,
        })

    prefix = os.path.join(out_dir, f'Art1_{test_name}')
    make_figure(test_name, param_data, stats_dict, prefix)
    save_csv(csv_rows, os.path.join(out_dir, f'Art1_{test_name}_stats.csv'))


def main():
    parser = argparse.ArgumentParser(
        description='Behavioural analysis — Article 1 (Control vs PTSD)')
    parser.add_argument('--test', choices=['OF', 'EPM', 'DLB', 'ALL'],
                        default='ALL', help='Which test to run')
    parser.add_argument('--no-outliers', action='store_true',
                        help='Skip IQR outlier removal')
    args = parser.parse_args()

    ts      = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    out_dir = os.path.join('results', ts)
    os.makedirs(out_dir, exist_ok=True)
    print(f"Output → {out_dir}/")

    tests = ['OF', 'EPM', 'DLB'] if args.test == 'ALL' else [args.test]
    for t in tests:
        run_test(t, out_dir, apply_iqr=not args.no_outliers)

    print(f"\nDone! All results → {out_dir}/")


if __name__ == '__main__':
    main()